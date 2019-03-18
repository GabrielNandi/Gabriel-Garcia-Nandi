# -*- coding: utf-8 -*-
"""
Created on Fri Feb 15 12:55:28 2019

@author: Luis Felipe
"""
import openpyxl as ex
import os
from pathlib import Path
import numpy as np
import teslamax
from teslamax import TeslaMaxGeometry, TeslaMaxPreDesign, TeslaMaxModel
from scipy.integrate import trapz
from pandas import Series, DataFrame

K_CRITICAL = 0.00075
B_rem = 1.26
B_min = 0.0
field_fraction = 0.3
TESLAMAX_PATH = Path.home() / "code" / "TeslaMax"
os.chdir(str(TESLAMAX_PATH))
PLAYGROUND_DIR = "teslamax-play"
RAMP_RESULTS_DIR_FMT = "ramp_Bh_%d_linha_%d"
teslamax_playground = TESLAMAX_PATH / PLAYGROUND_DIR


params_optimization_ref = {"h_fc": 0.025,
                "R_e": 0.8,
                "n_II": 0,
                "n_IV": 5,
                "phi_C_IV": 0.0,
                "phi_C_II": 0.0,
                "phi_S_II": 0.0,
                "phi_S_IV": 60.0,
                "mu_r_II": 1.05,
                "mu_r_IV": 1.05,
              "linear_iron": 1,
              "mu_r_iron": 5e3,
             }
n = int(params_optimization_ref["n_II"] + params_optimization_ref["n_IV"])
params_optimization_ref = teslamax.expand_parameters_from_remanence_array(
    B_rem*np.ones(n), 
    params_optimization_ref, 
    "B_rem")

params_optimization_ref["F_M[%]"] = field_fraction*100
params = params_optimization_ref.copy()



for i in range(70,99):
    wb = ex.load_workbook("Planejamento das simulações - Modelo Analítico - NOVO.xlsx")
    ws = wb['Simulações_Modelo']
    R_i = ws.cell(row=i, column=2).value*0.001
    R_o = ws.cell(row=i, column=3).value*0.001
    R_g = ws.cell(row=i, column=4).value*0.001
    B_max = ws.cell(row=i, column=5).value
    params["R_g"] = R_g
    params["R_o"] = R_o
    params["R_i"] = R_i
    target_args = (B_max, B_min, field_fraction)
    print ("SIMULACAO COM B_max IGUAL A %.3f" %(B_max))
    print ("SIMULACAO COM R1 IGUAL A %.1f" %(R_i*1000))
    print ("SIMULACAO COM R2 IGUAL A %.1f" %(R_o*1000))
    print ("SIMULACAO COM R3 IGUAL A %.2f" %(R_g*1000))
    #Varia o K crítico dependendo de B_max. Quando o B_max é alto, um K crítico menor é necessário. Porém, quanto menor o B_max, Às vezes ele não converge para um K_critico tão pequeno
    if B_max <= 1.3:
        if B_max <= 1.2:
            K_CRITICAL = 0.0009
        else:    
            K_CRITICAL = 0.0007
    else: 
        if B_max >= 1.5:
            K_CRITICAL = 0.0005
        else: K_CRITICAL = 0.0006
    K=1e3
    K1=1e4
    R_s = R_g
    dRs=10e-3
    RS1 = R_s
    K2 = [1e3]
    while (K > K_CRITICAL):
        #para otimizar o tempo de simulação, o incremento de Rs varia de acordo com o K anterior. Quanto mais próximo do K crítico, menor o incremento
        if K == 1e3:
            dRs = 5e-3
        else:
            if K >= 0.2:
                dRs = 0.025
            else:
                if K>= 0.1:
                    dRs = 0.02
                else:
                    if K >= 0.05:
                        dRs = 0.015
                    else:
                        if K>= 0.02:
                            dRs= 0.011
                        else: 
                            if K>= 0.01:
                                dRs = 0.008
                            else:
                                if K> 0.005:
                                    dRs = 0.006
                                else: 
                                    if K>0.002:
                                        dRs = 0.003
                                    else:
                                        if K > 0.0015:
                                            dRs = 0.002
                                        else: 
                                            if K > 0.0009:
                                                dRs = 0.001
                                            else: dRs = 0.0005
        #para caso haja descontinuidade OU que K comece a crescer (não achou K crítico), interrompe a simulação e aceita o menor K obtido
        if len(K2) > 4:
            if K2[-1] > K2[-2] and K2[-2] > K2[-3] and K2[-3] > K2[-4]:
                break
                #caso haja as comuns descontinuidades (K chegando continuamente próximo de K crítico e do nada 'pula' para um valor bem mais alto)
                if K2[-2] > 2*K2[-3] and K2[-1] > K2[-3]:
                    print ("descontinuidade")
                    break
        R_s = R_s + dRs
        print ("SIMULACAO COM R4 IGUAL A %.3f" %(R_s*1000))
        params["R_s"] = R_s
        if R_s >= 4*R_g:
            break
        else: 
            try:
                #calcula ângulos_rem ótimos e o respectivo K (função custo)
                print (target_args)
                tmpd = TeslaMaxPreDesign(params)
                target_function = teslamax.calculate_ramp_profile
                alpha_B_rem_g = tmpd.get_optimal_remanence_angles(target_function, target_args)
                print (alpha_B_rem_g)
                K = tmpd.calculate_functional_target(alpha_B_rem_g,
                                                      target_function,
                                                      target_args)
                print(K)
                #para caso a função não ache o K crítico e continue a simulação, ele salva o menor K obtido
                if K < K1:
                    K1 = K
                    RS1 = R_s
                K2.append(K)
                sdir = RAMP_RESULTS_DIR_FMT %(B_max*1e3, i)
                teslamax_playground = str(TESLAMAX_PATH / PLAYGROUND_DIR / sdir )
    
                tmm_g = TeslaMaxModel(tmpd,
                          alpha_B_rem_g,
                          path=teslamax_playground)
                tmm_g.run()
            except:
                print ("NAO SIMULOU")
                continue
           
    RS2 = RS1*1000
    #abre pasta teslamax_play para pegar arquivo com dados de B x phi
    sdir = RAMP_RESULTS_DIR_FMT %(B_max*1e3, i)
    teslamax_playground = TESLAMAX_PATH / PLAYGROUND_DIR / sdir
    profile_file = teslamax_playground / teslamax.MAGNETIC_PROFILE_FILENAME
    phi_vector_g, B_vector_g = teslamax.read_comsol_profile_data(str(profile_file))
    phi_vector_g = np.deg2rad(phi_vector_g)
    #Pegar B e phi apenas até a metade do primeiro quadrante (área de campo alto)
    B_vector_g_2 = B_vector_g[0:90]
    phi_vector_g_2 = phi_vector_g[0:90]
    #calcula a energia total nesta área
    W = ((R_g*1000)**2 - (R_o*1000)**2)*0.5*trapz(B_vector_g_2**2,phi_vector_g_2)
    print ('W')
    print(W)
    #bota na planilha do excel e salva
    ws.cell(row=i, column=18).value = "%.4f" %W
    ws.cell(row=i, column=12).value = "%.2f" %RS2
    ws.cell(row=i, column=13).value = "%.7f" %K1
    wb.save("Planejamento das simulações - Modelo Analítico - NOVO.xlsx")
