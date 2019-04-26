# -*- coding: utf-8 -*-
"""
Created on Thu Mar 28 09:39:51 2019

@author: Luis Felipe
"""

from pathlib import Path
import os
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import teslamax
from teslamax import TeslaMaxGeometry, TeslaMaxPreDesign, TeslaMaxModel
from scipy.integrate import trapz
import openpyxl as ex
import nemplot

B_HIGH_LIMIT = 1.85

TESLAMAX_PATH = Path.home() / "code" / "TeslaMax"

PLAYGROUND_DIR = "teslamax-play"

RAMP_RESULTS_DIR_FMT = "ramp_Bh_%d_FM_%d"
os.chdir(TESLAMAX_PATH)
os.getcwd()

def optimize_ramp(B_high,field_fraction,params):
    """
    Run the TeslaMax optimization procedure, trying to match the profile
    to a ramp profile between 0 and 'B_high' with 'field_fraction', using the
    parameters in the dictionary 'params'.
    
    Return a TeslaMaxModel object, representing the optimized magnet,
    and the value of the cost function. The TeslaMaxModel object is already populated with simulation results.
    
    The simulation results are stored in directories inside "teslamax-play".
    """
    
    tmpd = TeslaMaxPreDesign(params)
    
    target_function = teslamax.calculate_ramp_profile
    
    B_low = 0.0
    
    target_args = (B_high,B_low,field_fraction)

    alpha_B_rem_g = tmpd.get_optimal_remanence_angles(target_function,target_args)
    cost_function = tmpd.calculate_functional_target(alpha_B_rem_g,
                                                    target_function,
                                                    target_args)
    
    sdir = RAMP_RESULTS_DIR_FMT %(B_high*1e3, field_fraction*1e2) 
    
    teslamax_playground = str(TESLAMAX_PATH / PLAYGROUND_DIR / sdir )
    
    tmm_g = TeslaMaxModel(tmpd,
                          alpha_B_rem_g,
                          path=teslamax_playground)
    tmm_g.run()
    
    return (tmm_g,cost_function)
nemplot.set_dpi(1500)
nemplot.set_fontsize(22)
nemplot.set_latex_font("Palatino")
def refine_list(original_list, factor):
    """
    Return a new list, inserting more elements between the number in 'original_list'.
    
    Assumes 'original_list' is evenly-spaced. The spacing between each element is divided by 'factor'
    
    >>>refine_list([1,2,3],factor=2)
    [1.0,1.5,2.0,2.5,3.0]
    """
    
    # calculate the original spacing between elements and refine it
    d = original_list[1] - original_list[0]
    d_refined = d / factor
        
    max_value = max(original_list)
    min_value = min(original_list)
    
    # the number of elements is the number of divisions between the limit values, plus one aditional
    n_refined = ((max_value - min_value) / d_refined) + 1
        
    return np.linspace(min_value,max_value,int(n_refined))
def plot_optimized_ramp(B_high,field_fraction):
    """
    Plot a comparison between the magnetic profile from the TeslaMax simulations
    and a ramp profile between 0 and 'B_high', with 'field_fraction'. The TeslaMax
    results are read from the appropriate files inside PLAYGROUND_DIR.
    
    Save the figure with the appropriate naming inside the 'figures' directory.
    """
 

    sdir = RAMP_RESULTS_DIR_FMT %(B_high*1e3, field_fraction*1e2) 
    
    teslamax_playground = TESLAMAX_PATH / PLAYGROUND_DIR / sdir
    
    profile_file = teslamax_playground / teslamax.MAGNETIC_PROFILE_FILENAME
    
    profile_fig_g, profile_axes_g = nemplot.create_plot(xlabel=r'$\phi [\si{\degree}]$',
                                                       ylabel=r'$B [\si{\tesla}]$')
    
    phi_vector_g, B_vector_g = teslamax.read_comsol_profile_data(str(profile_file))
    
    lw = 2.0
    profile_axes_g.plot(phi_vector_g,B_vector_g,linewidth=lw)
    profile_axes_g.set_xlim(0,np.max(phi_vector_g))
    profile_axes_g.set_xticks(np.linspace(0,np.max(phi_vector_g),9))
    profile_axes_g.grid(True)
    profile_axes_g.set_ylim(0,B_HIGH_LIMIT)
    
    ticks = profile_axes_g.get_yticks()
    profile_axes_g.set_yticks(refine_list(ticks,4),minor=True)
        
    #profile_axes_g.set_yticks(np.arange(0,B_HIGH_LIMIT+0.2,0.2),minor=True)
    
    B_low = 0.0
    B_inst_vector = teslamax.calculate_ramp_profile(phi_vector_g,B_high,B_low,field_fraction)
    profile_axes_g.plot(phi_vector_g,B_inst_vector,'--',linewidth=lw)

    return (phi_vector_g , B_vector_g)
wb = ex.load_workbook("Planejamento das simulações - Modelo Analítico - NOVO.xlsx")
ws = wb['Simulações_Modelo']
# Fixed parameters



FIXED_PARAMETERS_MAGNET = {"h_fc": 0.040,
                "R_e": 0.8,
                "n_II": 0,
                "n_IV": 5,
                "phi_C_II": 0,
                "phi_C_IV": 0,
                "phi_S_II" : 0,
                "phi_S_IV": 60,          
                "mu_r_II": 1.05,
                "mu_r_IV": 1.05,
              "linear_iron": 1,
              "mu_r_iron": 5e3,
             }
B_rem = 1.26
field_fraction = 0.25
params_optimization_2 = FIXED_PARAMETERS_MAGNET.copy()
n_II = params_optimization_2["n_II"]
n_IV = params_optimization_2["n_IV"]
n = n_II + n_IV
params_optimization_2 = teslamax.expand_parameters_from_remanence_array(B_rem*np.ones(n), 
                                                                      params_optimization_2, 
                                                                      "B_rem")
print("Parameters:")
print(params_optimization_2)
print ("\n\n")
for i in [100]:
    #B_high da tabela
    params_optimization_2["R_i"] = float(ws.cell(row=i, column=2).value)*0.001
    params_optimization_2["R_o"] = float(ws.cell(row=i, column=3).value)*0.001
    params_optimization_2["R_g"] = float(ws.cell(row=i, column=4).value)*0.001
    params_optimization_2["R_s"] = float(ws.cell(row=i, column=12).value)*0.001
    B_high = ws.cell(row=i, column=5).value
    print ("Linha %d. R1 = %.1f, R2= %.1f, R3=%.2f, R4=%.1f, B_max=%.2f" %(i,params_optimization_2["R_i"]*1000,params_optimization_2["R_o"]*1000,params_optimization_2["R_g"]*1000,params_optimization_2["R_s"]*1000,B_high))
    tmm_g_2, K_2= optimize_ramp(B_high,field_fraction,params_optimization_2)
    print("Value of the objective function:")
    print(K_2)
    phi_vector_g, B_vector_g = plot_optimized_ramp(B_high,field_fraction)
    plt.show()
    
    #B_high proporcional ao B_rem
    #B_high = B_high*1.35/1.26
    B_high = 1.18
    print ("Linha %d. R1 = %.1f, R2= %.1f, R3=%.2f, R4=%.1f, B_max=%.2f" %(i,params_optimization_2["R_i"]*1000,params_optimization_2["R_o"]*1000,params_optimization_2["R_g"]*1000,params_optimization_2["R_s"]*1000,B_high))
    tmm_g_2, K_2= optimize_ramp(B_high,field_fraction,params_optimization_2)
    print("Value of the objective function:")
    print(K_2)
    phi_vector_g, B_vector_g = plot_optimized_ramp(B_high,field_fraction)
    plt.show()
    
    B_rem = 1.37
    params_optimization_2 = teslamax.expand_parameters_from_remanence_array(B_rem*np.ones(n), 
                                                                      params_optimization_2, 
                                                                      "B_rem")
    B_high = 1.18
    print ("Linha %d. R1 = %.1f, R2= %.1f, R3=%.2f, R4=%.1f, B_max=%.2f" %(i,params_optimization_2["R_i"]*1000,params_optimization_2["R_o"]*1000,params_optimization_2["R_g"]*1000,params_optimization_2["R_s"]*1000,B_high))
    tmm_g_2, K_2= optimize_ramp(B_high,field_fraction,params_optimization_2)
    print("Value of the objective function:")
    print(K_2)
    phi_vector_g, B_vector_g = plot_optimized_ramp(B_high,field_fraction)
    plt.show()
    
    B_high = 1.3
    print ("Linha %d. R1 = %.1f, R2= %.1f, R3=%.2f, R4=%.1f, B_max=%.2f" %(i,params_optimization_2["R_i"]*1000,params_optimization_2["R_o"]*1000,params_optimization_2["R_g"]*1000,params_optimization_2["R_s"]*1000,B_high))
    tmm_g_2, K_2= optimize_ramp(B_high,field_fraction,params_optimization_2)
    print("Value of the objective function:")
    print(K_2)
    phi_vector_g, B_vector_g = plot_optimized_ramp(B_high,field_fraction)
    plt.show()
    
